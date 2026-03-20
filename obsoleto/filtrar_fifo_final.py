import pandas as pd
import os
from pyxlsb import open_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import win32com.client


# Índice de la columna Mes/año (para formato fecha)
IDX_MES_ANIO = 59


def normalizar_serie(serie):
    """Normaliza serie quitando ceros iniciales para comparación"""
    if serie is None:
        return ''
    return str(serie).strip().lstrip('0')


def cargar_series_validadas(archivo_path):
    """Carga las series validadas en un set NORMALIZADO para búsqueda rápida O(1)"""
    df = pd.read_excel(archivo_path, dtype=str)
    series_raw = df['Numero_de_Serie'].dropna().astype(str).str.strip().tolist()

    # Crear set con series normalizadas (sin ceros iniciales)
    series_norm = {normalizar_serie(s) for s in series_raw}

    print(f"  Cargadas {len(series_raw)} series de {archivo_path}")
    print(f"  Series únicas normalizadas: {len(series_norm)}")
    return series_norm


def leer_xlsb_como_lista(archivo_path):
    """Lee un archivo xlsb y retorna cabeceras y filas (preservando números en Mes/año)"""
    print(f"  Leyendo {archivo_path}...")
    filas = []
    cabeceras = []

    with open_workbook(archivo_path) as wb:
        with wb.get_sheet(1) as sheet:
            for i, row in enumerate(sheet.rows()):
                fila = []
                for col_idx, cell in enumerate(row):
                    valor = cell.v
                    if valor is None:
                        fila.append('')
                    elif col_idx == IDX_MES_ANIO and isinstance(valor, (int, float)):
                        # Preservar número de fecha Excel para columna Mes/año
                        fila.append(valor)
                    elif isinstance(valor, float):
                        # Convertir float a string sin notación científica
                        if valor == int(valor):
                            fila.append(str(int(valor)))
                        else:
                            fila.append(f"{valor:.10f}".rstrip('0').rstrip('.'))
                    else:
                        fila.append(str(valor))

                if i == 0:
                    cabeceras = fila
                else:
                    filas.append(fila)

    print(f"  Leídas {len(filas)} filas")
    return cabeceras, filas


def normalizar_almacen(valor):
    """Normaliza el valor de Almacen Agrupado (reemplaza non-breaking space)"""
    if valor is None:
        return ''
    return str(valor).replace('\xa0', ' ').strip()


def guardar_excel_como_texto(cabeceras, filas, archivo_salida):
    """Guarda los datos en Excel con celdas como texto (excepto Mes/año como número)"""
    print(f"  Guardando {len(filas)} filas en {archivo_salida}...")

    wb = Workbook()
    ws = wb.active

    # Escribir cabeceras
    for col_idx, header in enumerate(cabeceras, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.number_format = '@'

    # Escribir datos
    col_mes_excel = IDX_MES_ANIO + 1  # Columna Excel (1-based)
    for row_idx, fila in enumerate(filas, 2):
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            if col_idx == col_mes_excel and isinstance(valor, (int, float)):
                # Columna Mes/año: formato fecha mmm-yy
                cell.number_format = 'mmm-yy'
            else:
                cell.number_format = '@'  # Formato texto

    wb.save(archivo_salida)
    print(f"  Archivo xlsx temporal guardado")


def convertir_xlsx_a_xlsb(ruta_xlsx, ruta_xlsb):
    """Convierte un archivo xlsx a xlsb usando Excel COM automation"""
    print(f"  Convirtiendo a XLSB...")

    # Convertir rutas a absolutas con backslashes (requerido por Excel COM)
    ruta_xlsx_abs = os.path.abspath(ruta_xlsx).replace('/', '\\')
    ruta_xlsb_abs = os.path.abspath(ruta_xlsb).replace('/', '\\')

    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(ruta_xlsx_abs)
        # Guardar como xlsb (FileFormat=50)
        wb.SaveAs(ruta_xlsb_abs, FileFormat=50)
        wb.Close()
        print(f"  Archivo XLSB guardado: {ruta_xlsb}")
    finally:
        excel.Quit()

    # Eliminar archivo xlsx temporal
    if os.path.exists(ruta_xlsx):
        os.remove(ruta_xlsx)
        print(f"  Archivo temporal xlsx eliminado")


def procesar_fija():
    """Procesa el archivo FIJA: CD VES (sin filtrar) + ALMACENES (validados)"""
    print("\n" + "="*60)
    print("PROCESANDO FIJA")
    print("="*60)

    # Cargar series validadas
    print("\n[1] Cargando series validadas...")
    series_validas = cargar_series_validadas("C:/FIFO/series_validadas_fija.xlsx")

    # Leer archivo xlsb
    print("\n[2] Leyendo archivo xlsb...")
    cabeceras, filas = leer_xlsb_como_lista("C:/FIFO/fifo_en_proceso_fija.xlsb")

    # Índices de columnas
    idx_almacen = 58  # Almacen Agrupado
    idx_serie = 3     # Serie

    # Separar CD VES (sin filtrar) y ALMACENES (para validar)
    print("\n[3] Separando CD VES y ALMACENES...")
    filas_cd_ves = []
    filas_almacenes = []
    for fila in filas:
        if len(fila) > idx_almacen:
            almacen = normalizar_almacen(fila[idx_almacen])
            if almacen == 'CD VES':
                filas_cd_ves.append(fila)
            elif almacen == 'ALMACENES':
                filas_almacenes.append(fila)

    print(f"  Filas CD VES (sin filtrar): {len(filas_cd_ves)}")
    print(f"  Filas ALMACENES (para validar): {len(filas_almacenes)}")

    # Validar series de ALMACENES (usando normalización - quitar ceros iniciales)
    print("\n[4] Validando series de ALMACENES...")
    filas_almacenes_validadas = []
    for fila in filas_almacenes:
        if len(fila) > idx_serie:
            serie = str(fila[idx_serie]).strip()
            serie_norm = normalizar_serie(serie)
            if serie_norm in series_validas:
                filas_almacenes_validadas.append(fila)

    print(f"  Filas ALMACENES validadas: {len(filas_almacenes_validadas)}")

    # Combinar: CD VES + ALMACENES validadas
    print("\n[5] Combinando CD VES + ALMACENES validadas...")
    filas_finales = filas_cd_ves + filas_almacenes_validadas
    print(f"  Total filas finales: {len(filas_finales)}")

    # Guardar como xlsx temporal y convertir a xlsb
    print("\n[6] Guardando archivo final...")
    xlsx_temp = "C:/FIFO/fifo_final_fija_temp.xlsx"
    xlsb_final = "C:/FIFO/fifo_final_fija.xlsb"
    guardar_excel_como_texto(cabeceras, filas_finales, xlsx_temp)
    convertir_xlsx_a_xlsb(xlsx_temp, xlsb_final)

    return len(filas), len(filas_cd_ves), len(filas_almacenes), len(filas_almacenes_validadas), len(filas_finales)


def procesar_movil():
    """Procesa el archivo MOVIL: CD VES (sin filtrar) + otros canales (validados)"""
    print("\n" + "="*60)
    print("PROCESANDO MOVIL")
    print("="*60)

    # Valores de canales a validar (normalizados)
    almacenes_a_validar = {
        'PDV',
        'Televentas VES',
        'Corporativo Ves',
        'Tienda Virtual VES'
    }

    # Cargar series validadas
    print("\n[1] Cargando series validadas...")
    series_validas = cargar_series_validadas("C:/FIFO/series_validadas_movil.xlsx")

    # Leer archivo xlsb
    print("\n[2] Leyendo archivo xlsb...")
    cabeceras, filas = leer_xlsb_como_lista("C:/FIFO/fifo_en_proceso_movil.xlsb")

    # Índices de columnas
    idx_almacen = 58  # Almacen Agrupado
    idx_serie = 3     # Serie

    # Separar CD VES (sin filtrar) y otros canales (para validar)
    print("\n[3] Separando CD VES y otros canales...")
    print(f"  Canales a validar: {almacenes_a_validar}")
    filas_cd_ves = []
    filas_otros = []
    for fila in filas:
        if len(fila) > idx_almacen:
            almacen = normalizar_almacen(fila[idx_almacen])
            if almacen == 'CD VES':
                filas_cd_ves.append(fila)
            elif almacen in almacenes_a_validar:
                filas_otros.append(fila)

    print(f"  Filas CD VES (sin filtrar): {len(filas_cd_ves)}")
    print(f"  Filas otros canales (para validar): {len(filas_otros)}")

    # Validar series de otros canales (usando normalización - quitar ceros iniciales)
    print("\n[4] Validando series de otros canales...")
    filas_otros_validadas = []
    for fila in filas_otros:
        if len(fila) > idx_serie:
            serie = str(fila[idx_serie]).strip()
            serie_norm = normalizar_serie(serie)
            if serie_norm in series_validas:
                filas_otros_validadas.append(fila)

    print(f"  Filas otros canales validadas: {len(filas_otros_validadas)}")

    # Combinar: CD VES + otros canales validados
    print("\n[5] Combinando CD VES + otros canales validados...")
    filas_finales = filas_cd_ves + filas_otros_validadas
    print(f"  Total filas finales: {len(filas_finales)}")

    # Guardar como xlsx temporal y convertir a xlsb
    print("\n[6] Guardando archivo final...")
    xlsx_temp = "C:/FIFO/fifo_final_movil_temp.xlsx"
    xlsb_final = "C:/FIFO/fifo_final_movil.xlsb"
    guardar_excel_como_texto(cabeceras, filas_finales, xlsx_temp)
    convertir_xlsx_a_xlsb(xlsx_temp, xlsb_final)

    return len(filas), len(filas_cd_ves), len(filas_otros), len(filas_otros_validadas), len(filas_finales)


def main():
    print("="*60)
    print("FILTRAR Y VALIDAR SERIES FIFO")
    print("="*60)

    # Procesar FIJA
    total_fija, cd_ves_fija, almacenes_fija, almacenes_val_fija, final_fija = procesar_fija()

    # Procesar MOVIL
    total_movil, cd_ves_movil, otros_movil, otros_val_movil, final_movil = procesar_movil()

    # Resumen final
    print("\n" + "="*60)
    print("RESUMEN FINAL")
    print("="*60)
    print(f"\nFIJA:")
    print(f"  Total filas originales:      {total_fija:,}")
    print(f"  CD VES (sin filtrar):        {cd_ves_fija:,}")
    print(f"  ALMACENES (antes validar):   {almacenes_fija:,}")
    print(f"  ALMACENES (validadas):       {almacenes_val_fija:,}")
    print(f"  TOTAL FINAL:                 {final_fija:,}")

    print(f"\nMOVIL:")
    print(f"  Total filas originales:      {total_movil:,}")
    print(f"  CD VES (sin filtrar):        {cd_ves_movil:,}")
    print(f"  Otros canales (antes val.):  {otros_movil:,}")
    print(f"  Otros canales (validadas):   {otros_val_movil:,}")
    print(f"  TOTAL FINAL:                 {final_movil:,}")

    print(f"\nArchivos generados:")
    print(f"  - C:/FIFO/fifo_final_fija.xlsb")
    print(f"  - C:/FIFO/fifo_final_movil.xlsb")


if __name__ == "__main__":
    main()
